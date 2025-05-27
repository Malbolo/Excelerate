import pandas as pd
from openpyxl import load_workbook

def insert_df_to_excel(df: pd.DataFrame,
                       input_path: str,
                       output_path: str = None,
                       sheet_name: str = None,
                       start_row: int = 1,
                       start_col: int = 1):
    """
    기존 Excel 파일에 df를 특정 위치에 삽입하고 저장합니다.
    
    Parameters:
    - df: 삽입할 pandas.DataFrame
    - input_path: 원본 엑셀 파일 경로(.xlsx, .xlsm)
    - output_path: 저장할 경로. None이면 input_path에 덮어쓰기
    - sheet_name: None이면 active sheet
    - start_row: 1부터 시작하는 삽입 시작 행 (기본 1)
    - start_col: 1부터 시작하는 삽입 시작 열 (기본 1 → A열)
    """
    # 1) 워크북 로드 (매크로 보존이 필요하면 keep_vba=True)
    wb = load_workbook(filename=input_path, keep_vba=False)
    ws = wb[sheet_name] if sheet_name else wb.active

    # 병합 셀 범위 리스트
    merged_ranges = list(ws.merged_cells.ranges)

    # 2) 헤더 삽입
    for j, col_name in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=j, value=col_name)

    # 3) 데이터 삽입
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, value in enumerate(row, start=start_col):
            # 병합 셀 내 포함 여부 확인: numeric bounds 검사
            target = None
            for merged in merged_ranges:
                if merged.min_row <= i <= merged.max_row and merged.min_col <= j <= merged.max_col:
                    target = (merged.min_row, merged.min_col)
                    break
            if target:
                ws.cell(row=target[0], column=target[1], value=value)
            else:
                ws.cell(row=i, column=j, value=value)

    # 4) 파일 저장
    save_path = output_path or input_path
    wb.save(save_path)
    print(f"Saved to: {save_path}")