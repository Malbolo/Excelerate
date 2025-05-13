# app/api/v1/endpoints/templates.py

import os
import urllib.parse
import uuid
import tempfile
from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.responses import JSONResponse, FileResponse
from app.utils.depend import get_minio_client
from app.utils.minio_client import MinioClient

# Linux libreoffice 용
import shutil
import subprocess
import glob
import platform

# Windows 테스트 용용
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from fastapi.concurrency import run_in_threadpool

router = APIRouter()

@router.post("/", summary="새 엑셀 템플릿 업로드")
async def upload_template(
    template_name: str = Form(...),
    file: UploadFile = File(...),
    minio: MinioClient = Depends(get_minio_client),
):
    # --- 0) 파일 확장자/콘텐츠 타입 검사 ---
    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext != ".xlsx":
        raise HTTPException(
            status_code=400,
            detail="지원되지 않는 파일 형식입니다. ‘.xlsx’ 파일만 업로드 가능합니다."
        )

    # 추가로 헤더 레벨에서 content_type 검사
    if file.content_type != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        raise HTTPException(
            status_code=400,
            detail=f"잘못된 콘텐츠 타입입니다: {file.content_type}"
        )

    # 1) 중복 이름 처리 (중복 시 자동 번호 부여)
    base_name = template_name
    counter = 1
    object_name = f"templates/{template_name}.xlsx"
    while minio.exists(object_name):
        template_name = f"{base_name} ({counter})"
        object_name = f"templates/{template_name}.xlsx"
        counter += 1

    # 2) OS에 맞는 임시 디렉터리 생성
    with tempfile.TemporaryDirectory() as tmpdir:
        tmp_path = os.path.join(tmpdir, f"{uuid.uuid4().hex}{ext}")

        # 3) 업로드된 파일을 임시 디렉터리에 저장
        try:
            contents = await file.read()
            with open(tmp_path, "wb") as f:
                f.write(contents)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"파일 저장 실패: {e}")

        # 4) MinIO에 템플릿 업로드
        try:
            minio.upload_template(template_name, tmp_path)
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"템플릿 업로드 실패: {e}")

    # with 블록 벗어나면 tmpdir 전체가 자동 삭제됩니다
    return JSONResponse(status_code=201, content={"result" : "success", "data" : {"message": f"'{template_name}' 업로드 완료."}})

@router.get("/", summary="등록된 템플릿 목록 조회")
async def list_templates(
    minio: MinioClient = Depends(get_minio_client),
):
    try:
        templates = minio.list_templates()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"템플릿 조회 실패: {e}")
    return JSONResponse(status_code=200, content={"result" : "success", "data" : {"templates": templates}})

@router.delete("/{template_name}", summary="특정 템플릿 삭제")
async def delete_template(
    template_name: str,
    minio: MinioClient = Depends(get_minio_client),
):
    try:
        minio.delete_template(template_name)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"템플릿 삭제 실패: {e}")
    return JSONResponse(status_code=200, content={"result" : "success", "data" : {"message": f"'{template_name}' 삭제 완료."}})

@router.get("/{template_name}/preview", summary="템플릿 첫 시트 미리보기 이미지")
async def preview_template(
    template_name: str,
    background_tasks: BackgroundTasks,
    minio: MinioClient = Depends(get_minio_client),
):
    # 1) 임시 디렉터리 생성
    tmpdir = tempfile.mkdtemp()
    xlsx_path = os.path.join(tmpdir, f"{template_name}.xlsx")
    png_path  = os.path.join(tmpdir, f"{template_name}.png")

    # 2) 다운로드
    try:
        minio.download_template(template_name, xlsx_path)
    except Exception as e:
        shutil.rmtree(tmpdir, ignore_errors=True)
        raise HTTPException(404, detail=f"템플릿을 찾을 수 없습니다: {e}")

    # 3) 시트 범위 계산
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    max_row = min(ws.max_row or 1, 40)
    max_col = min(ws.max_column or 1, 20)
    end_col = get_column_letter(max_col)   # 20 → "T"
    cell_range = f"A1:{end_col}{max_row}"  # "A1:T40"

    # 4) 플랫폼별 이미지 변환
    system = platform.system()
    if system == "Windows":
        # COM 기반 excel2img (로컬 테스트용). linux에선 import 시도 시 에러 뜨므로 분리
        import pythoncom
        import excel2img
        def do_export_com():
            pythoncom.CoInitialize()
            excel2img.export_img(xlsx_path, png_path, ws.title, cell_range)
        try:
            await run_in_threadpool(do_export_com)
        except Exception as e:
            shutil.rmtree(tmpdir, ignore_errors=True)
            raise HTTPException(500, detail=f"Windows COM 변환 실패: {e}")

    else:
        # LibreOffice headless (EC2/Linux)
        soffice = shutil.which("soffice") or shutil.which("libreoffice")
        if not soffice:
            shutil.rmtree(tmpdir, ignore_errors=True)
            raise HTTPException(500, detail="LibreOffice가 설치되어 있지 않습니다.")

        ws.print_area = cell_range # 범위를 출력 범위로 설정
        wb.save(xlsx_path) # 변경 사항 저장

        cmd = [
            soffice,
            "--headless",
            "--convert-to", "pdf:calc_pdf_Export",
            "--outdir", tmpdir,
            xlsx_path
        ]
        try:
            subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        except subprocess.CalledProcessError as e:
            stderr = e.stderr.decode(errors="ignore")
            shutil.rmtree(tmpdir, ignore_errors=True)
            raise HTTPException(500, detail=f"LibreOffice 변환 실패: {stderr}")

        # ② PDF→PNG 변환 (ImageMagick 필요)
        pdf_path = os.path.join(tmpdir, f"{template_name}.pdf")
        png_base = os.path.join(tmpdir, template_name)

        # 4-2) PDF → PNG (첫 페이지만)
        try:
            subprocess.run(
                [
                    "pdftoppm",
                    "-singlefile",    # 단일 PNG 파일
                    "-png",
                    "-f", "1",        # 첫 페이지
                    "-scale-to-x", "1024",
                    pdf_path,
                    png_base
                ],
                check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE
            )
        except subprocess.CalledProcessError as e:
            shutil.rmtree(tmpdir, ignore_errors=True)
            detail = e.stderr.decode(errors="ignore")
            raise HTTPException(500, detail=f"PDF→PNG 변환 실패: {detail}")
        
        # 5) 생성된 PNG 파일 찾기
        png_files = glob.glob(os.path.join(tmpdir, "*.png"))
        if not png_files:
            shutil.rmtree(tmpdir, ignore_errors=True)
            raise HTTPException(500, detail="변환된 PNG 파일을 찾을 수 없습니다.")
        png_path = png_files[0]

    # 5) 응답 후 임시 디렉터리 삭제 스케줄
    background_tasks.add_task(shutil.rmtree, tmpdir, True)

    # 6) 반환
    response = FileResponse(
        png_path,
        media_type="image/png",
    )

    # RFC5987 포맷으로 한글 이름을 UTF-8 URL 인코딩
    filename = os.path.basename(png_path)  # e.g. "테스트.png"
    quoted_fname = urllib.parse.quote(filename, safe="")  

    # 다운로드가 아닌 인라인 뷰로
    response.headers["Content-Disposition"] = f"inline; filename*=UTF-8''{quoted_fname}"
    return response