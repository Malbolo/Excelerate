
import re
import traceback
import json
from datetime import datetime
from pydantic import BaseModel
from langchain_core.prompts import ChatPromptTemplate, SystemMessagePromptTemplate, HumanMessagePromptTemplate, AIMessagePromptTemplate
from difflib import SequenceMatcher
from typing import Optional, List, Tuple
import pandas as pd
from openpyxl import load_workbook

def extract_relevant_code_fuzzy(
    original_code: str,
    commands: List[str],
    similarity_threshold: float = 0.6
) -> Optional[str]:
    """
    original_code 에서 commands 각각에 대해
    1) ‘# …’ 주석 라인을 찾아서,
    2) 주석 텍스트와 커맨드를 양쪽 모두 구두점·공백 제거 후 비교하여
       ratio >= similarity_threshold 인 블록을
    3) # 주석 바로 다음 줄부터 다음 주석/빈 줄 전까지 스니펫으로 추출,
    4) 여러 커맨드에 해당하는 스니펫이 있으면 \n\n 으로 합쳐서 리턴합니다.
    """
    lines = original_code.splitlines()
    snippets: List[str] = []

    def clean_text(s: str) -> str:
        # "#" 빼고, 마침표/따옴표/백틱 등 제거, 양쪽 공백 strip
        return re.sub(r"[\.\"'`]", "", s).lstrip("#").strip()

    # 주석 라인만 모아뒀다가, comment_map[idx] = clean_comment
    comment_map = {
        idx: clean_text(line)
        for idx, line in enumerate(lines)
        if line.strip().startswith("#")
    }

    for cmd in commands:
        cmd_clean = re.sub(r"[\.\"'`]", "", cmd).strip()
        # 1) 이 커맨드와 주석 중 best match 찾기
        best_idx = None
        best_ratio = 0.0
        for idx, comment in comment_map.items():
            ratio = SequenceMatcher(None, comment, cmd_clean).ratio()
            if ratio > best_ratio:
                best_ratio = ratio
                best_idx = idx

        # 2) 임계치 미달 무시
        if best_idx is None or best_ratio < similarity_threshold:
            continue

        # 3) best_idx + 1 부터 다음 주석/빈 줄 전까지 스니펫 추출
        snippet_lines = [lines[best_idx]]
        for body in lines[best_idx + 1:]:
            if body.strip().startswith("#") or body.strip() == "":
                break
            snippet_lines.append(body)
        if snippet_lines:
            snippets.append("\n".join(snippet_lines))

    if not snippets:
        return None

    # 4) 여러 개면 빈 줄 두 개로 구분해서 합침
    return "\n\n".join(snippets)

## 추 후 다양한 기능을(단순 불러오기, dataframe 치환, 다중 저장 등) 수행시키고 싶다면 해당 템플릿을 사용해 엑셀 코드를 생성하도록 수정해야 할 것
def make_excel_template() -> ChatPromptTemplate:
    """
    Airflow DAG Task에서 실행할 수 있는 엑셀 코드 스니펫을 생성하는 Prompt
    - 작업 디렉토리는 mkdtemp()로 생성
    - 자동 삭제 로직은 포함하지 않음 (downstream Task에서 처리)
    - 다운로드, DataFrame 삽입, 업로드만 수행
    - few-shot 예시 2개 포함
    """
    return ChatPromptTemplate.from_messages([
        # 시스템 메시지: 역할 정의
        SystemMessagePromptTemplate.from_template(
            """
당신은 MinioClient에 저장된 엑셀 템플릿 파일을 불러와 Dataframe을 붙여넣는 Python 코드를 생성하는 전문가입니다.
- 작업 디렉토리는 mkdtemp()로 만들고 자동 삭제 코드는 작성하지 마세요.
- 기존 템플릿 파일 다운로드 → DataFrame 삽입 → 결과 업로드 로직만 포함해주세요.
- insert_df_to_excel과 minio = MinioClient()를 사용하여 코드를 완성하세요.
- import문이나 함수 정의(`def …`)나 `return` 문은 쓰지 마세요.
- ```로 감싸지 말고 코드를 작성해 주세요.
- 코드 위, 아래에 `# Excel 작업시작`, `# Excel 작업 끝`이라는 주석을 달아주세요.
- 사용자 요청을 코드 최상단에 주석으로 표시해주세요.
"""
        ),

        # 첫 번째 페어: 기본 예시
        HumanMessagePromptTemplate.from_template(
            """
템플릿 EOE 의 2열 2행에 DataFrame을 붙여넣고, 결과를 out1으로 저장 후 업로드해주세요.
"""
        ),
        AIMessagePromptTemplate.from_template(
            """
# 템플릿 EOE 의 2열 2행에 DataFrame을 붙여넣고, 결과를 out1으로 저장 후 업로드해주세요.
# Excel 작업 시작
workdir = mkdtemp()
tpl    = os.path.join(workdir, "EOE.xlsx")
out    = os.path.join(workdir, "out1.xlsx")
minio = MinioClient()
minio.download_template("EOE", tpl)
insert_df_to_excel(
    df, tpl, out,
    sheet_name=None,
    start_row=2,
    start_col=2,
)
minio.upload_result("auto", "out1.xlsx", out)
# Excel 작업 끝
"""
        ),

#         # 두 번째 페어: C5 예시
#         HumanMessagePromptTemplate.from_template(
#             """
# report 템플릿의 C5 위치에 DataFrame을 삽입 후 동일 이름으로 업로드해주세요.
# """
#         ),
#         AIMessagePromptTemplate.from_template(
#             """
# # report 템플릿의 C5 위치에 DataFrame을 삽입 후 동일 이름으로 업로드해주세요.
# # Excel 작업 시작
# workdir = mkdtemp()
# tpl    = os.path.join(workdir, "report.xlsx")
# out    = os.path.join(workdir, "report.xlsx")
# minio = MinioClient()
# minio.download_template("report.xlsx", tpl)
# insert_df_to_excel(
#     df, tpl, out,
#     sheet_name=None,
#     start_row=5,
#     start_col=3,
# )
# minio.upload_result("auto", "report.xlsx", out)
# # Excel 작업 끝
# """
#         ),

        # 실제 사용자 요청
        HumanMessagePromptTemplate.from_template(
            "{input}"
        ),
    ])

def extract_error_info(exc: Exception, code_body: str, stage: str, commands: list[str], similarity_threshold: float = 0.5) -> dict:
    """
    Exception과 원본 코드 문자열, 실패 단계를 받아서
    빈 줄을 블록 구분자로 사용해, 해당 블록 시작(가장 가까운 빈 줄 이후)부터
    에러 라인까지의 snippet과, 블록 시작 주석을 command로 추출합니다.
    그 command가 commands 내에서 얼마나 유사한지 보고, similarity_threshold 이상이면 그 인덱스
    를 반환합니다.
    """
    # 1) traceback에서 마지막 프레임 정보 추출
    tb_last = traceback.extract_tb(exc.__traceback__)[-1]
    raw_lineno = tb_last.lineno

    # 2) 코드 본문을 줄 단위로 분리
    lines = code_body.splitlines()
    total = len(lines)

    # 3) 1-based lineno → 0-based index로 변환, 범위 clamp
    err_idx = min(raw_lineno - 1, total - 1)

    # 4) 블록 시작 찾기: err_idx 바로 위부터 검색해 최초의 빈 줄 이후를 블록 시작으로
    block_start: int = 0
    for idx in range(err_idx, -1, -1):
        line_strip = lines[idx].lstrip()
        # (1) 주석으로 시작하고, (2) 바로 위가 빈 줄이거나 맨 앞이면 블록 시작
        if line_strip.startswith("#") and (idx == 0 or lines[idx-1].strip() == ""):
            block_start = idx
            break

    # 5) 블록 시작 줄(주석)에서 command 추출
    #    예: "# defect_rate가 1 이상인 데이터만 필터링 해주세요."
    raw_comment = lines[block_start].lstrip()
    command = raw_comment[1:].strip() if raw_comment.startswith("#") else ""
    
    # 6) command_index 찾기
    best_idx   : Optional[int]   = None
    best_ratio : float = 0.0
    for idx, cand in enumerate(commands):
        # SequenceMatcher.ratio()는 0~1 사이 값을 반환
        ratio = SequenceMatcher(None, command, cand).ratio()
        if ratio > best_ratio:
            best_ratio, best_idx = ratio, idx

    if best_ratio >= similarity_threshold:
        command_index = best_idx
    else:
        command_index = None

    # 7) snippet 생성 (block_start 부터 err_idx 까지)
    snippet_lines = []
    for i in range(block_start, err_idx + 1):
        text = lines[i].rstrip()
        prefix = "→" if i == err_idx else "  "
        snippet_lines.append(f"{prefix} {i+1:4d}: {text}")
    snippet = "\n".join(snippet_lines)

    return {
        "error_msg": {
            "stage":    stage,
            "message":  str(exc),
            "file":     tb_last.filename,
            "line":     raw_lineno,
            "command":  command,
            "command_index": command_index,
            "code":     snippet,
        }
    }

def log_filter(entry: BaseModel) -> str:
    # Pydantic 객체 → dict
    entry_dict = entry.model_dump()
    msgs = entry_dict.get("input", [])

    # 첫 번째 system 메시지
    first_sys = next((m for m in msgs if m.get("role") == "system"), None)
    # 마지막 human 메시지
    last_hum  = next((m for m in reversed(msgs) if m.get("role") == "human"), None)

    # input 필드에 두 메시지만 남기기
    entry_dict["input"] = [m for m in (first_sys, last_hum) if m]

    # datetime 직렬화 처리용 default 함수
    def _json_default(o):
        if isinstance(o, datetime):
            return o.isoformat()
        raise TypeError(f"{o!r} is not JSON serializable")

    # JSON 문자열로 반환 (한글 깨짐 방지, datetime → ISO 포맷)
    return json.dumps(entry_dict, ensure_ascii=False, default=_json_default)

def insert_df_to_excel(df: pd.DataFrame,
                       input_path: str,
                       output_path: str = None,
                       sheet_name: str = None,
                       start_row: int = 5,
                       start_col: int = 1):
    """
    기존 Excel 파일에 df를 특정 위치에 삽입하고 저장합니다.
    
    Parameters:
    - df: 삽입할 pandas.DataFrame
    - input_path: 원본 엑셀 파일 경로(.xlsx, .xlsm)
    - output_path: 저장할 경로. None이면 input_path에 덮어쓰기
    - sheet_name: None이면 active sheet
    - start_row: 1부터 시작하는 삽입 시작 행 (기본 5)
    - start_col: 1부터 시작하는 삽입 시작 열 (기본 1 → A열)
    """
    # 1) 워크북 로드 (매크로 보존이 필요하면 keep_vba=True)
    wb = load_workbook(filename=input_path, keep_vba=False)
    ws = wb[sheet_name] if sheet_name else wb.active

    # 병합 셀 범위 리스트
    merged_ranges = list(ws.merged_cells.ranges)

    # 2) 헤더 삽입
    for j, col_name in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=j, value=col_name)

    # 3) 데이터 삽입
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, value in enumerate(row, start=start_col):
            # 병합 셀 내 포함 여부 확인: numeric bounds 검사
            target = None
            for merged in merged_ranges:
                if merged.min_row <= i <= merged.max_row and merged.min_col <= j <= merged.max_col:
                    target = (merged.min_row, merged.min_col)
                    break
            if target:
                ws.cell(row=target[0], column=target[1], value=value)
            else:
                ws.cell(row=i, column=j, value=value)

    # 4) 파일 저장
    save_path = output_path or input_path
    wb.save(save_path)
    print(f"Saved to: {save_path}")

def make_excel_code_snippet(template_name: str,
                             output_name: str,
                             start_row: int,
                             start_col: int,
                             sheet_name: str = None,
                             user_id: str = "test") -> str:
    """
    Excel 작업을 Airflow 등에서 재실행할 수 있도록 하는 Python 코드 스니펫을 생성합니다.
    """
    sheet_repr = repr(sheet_name) if sheet_name is not None else 'None'
    return f"""
# Excel 작업 시작
# {template_name} 템플릿의 {start_col}열 {start_row}행에 dataframe 붙여넣기
workdir = mkdtemp()
tpl    = os.path.join(workdir, "{template_name}.xlsx")
out    = os.path.join(workdir, "{output_name}.xlsx")

# 다운로드
minio = MinioClient()
minio.download_template("{template_name}", tpl)

# 삽입
insert_df_to_excel(
    df, tpl, out,
    sheet_name={sheet_repr},
    start_row={start_row},
    start_col={start_col},
)

# 업로드
minio.upload_result("{user_id}", "{template_name}", out)
# Excel 작업 끝
"""