import os
import uuid
import tempfile
import shutil
import subprocess
import glob
import platform

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from PIL import Image, ImageChops

from fastapi import HTTPException
from fastapi.concurrency import run_in_threadpool
from app.utils.minio_client import MinioClient

class TemplateService:
    def __init__(self, minio: MinioClient):
        self.minio = minio

    def upload_template(self, template_name: str, file_bytes: bytes, filename: str, content_type: str)-> str:
        # 0) 파일 확장자/콘텐츠 타입 검사
        ext = os.path.splitext(filename or "")[1].lower()
        if ext != ".xlsx":
            raise HTTPException(
                status_code=400,
                detail="지원되지 않는 파일 형식입니다. ‘.xlsx’ 파일만 업로드 가능합니다."
            )
        if content_type != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            raise HTTPException(status_code=400, detail=f"잘못된 콘텐츠 타입입니다: {content_type}")

        # 1) 중복 이름 처리
        base = template_name
        counter = 1
        object_name = f"templates/{template_name}.xlsx"
        while self.minio.exists(object_name):
            template_name = f"{base} ({counter})"
            object_name = f"templates/{template_name}.xlsx"
            counter += 1

        # 2) 임시 파일 생성 및 업로드
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp_path = os.path.join(tmpdir, f"{uuid.uuid4().hex}{ext}")
            try:
                with open(tmp_path, "wb") as f:
                    f.write(file_bytes)
                self.minio.upload_template(template_name, tmp_path)
            except Exception as e:
                raise HTTPException(status_code=500, detail=f"템플릿 업로드 실패: {e}")

        return template_name

    async def generate_preview(self, template_name: str, tmpdir: str) -> str:
        xlsx_path = os.path.join(tmpdir, f"{template_name}.xlsx")
        png_path = os.path.join(tmpdir, f"{template_name}.png")

        # 1) 다운로드
        try:
            self.minio.download_template(template_name, xlsx_path)
        except Exception as e:
            raise HTTPException(404, detail=f"템플릿을 찾을 수 없습니다: {e}")

        # 2) 셀 범위 계산
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb.active
        max_row = min(ws.max_row or 1, 40)
        max_col = min(ws.max_column or 1, 20)
        end_col = get_column_letter(max_col)
        cell_range = f"A1:{end_col}{max_row}"

        # 3) 플랫폼별 변환
        system = platform.system()
        if system == "Windows":
            # COM 기반 excel2img (로컬 테스트용). linux에선 import 시도 시 에러 뜨므로 분리
            import pythoncom, excel2img
            def do_export():
                pythoncom.CoInitialize()
                excel2img.export_img(xlsx_path, png_path, ws.title, cell_range)
            try:
                await run_in_threadpool(do_export)
            except Exception as e:
                raise HTTPException(500, detail=f"Windows COM 변환 실패: {e}")
        else:
            # LibreOffice headless (EC2/Linux)
            soffice = shutil.which("soffice") or shutil.which("libreoffice")
            if not soffice:
                raise HTTPException(500, detail="LibreOffice가 설치되어 있지 않습니다.")
            
            ws.print_area = cell_range # 범위를 출력 범위로 설정
            ws.page_setup.paperSize = ws.PAPERSIZE_A3 # 더 큰 용지 사용
            ws.page_setup.scale = 75 # 75% 축소해서 다 담기게끔 설정
            ws.page_setup.fitToWidth = 1 # 가로 폭에 맞추기
            ws.page_margins = PageMargins(left=0.1, right=0.1, top=0.1, bottom=0.1) # PDF화의 여백 좁게 설정
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE # 가로방향 출력
            wb.save(xlsx_path) # 변경 사항 저장

            # xlsx→PDF 변환환
            cmd = [
                soffice,
                "--headless",
                "--convert-to", "pdf:calc_pdf_Export",
                "--outdir", tmpdir,
                xlsx_path
            ]
            try:
                subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            except subprocess.CalledProcessError as e:
                detail = e.stderr.decode(errors="ignore")
                raise HTTPException(500, detail=f"LibreOffice 변환 실패: {detail}")

            # PDF→PNG 변환 (poppler-uilts 사용)
            pdf_path = os.path.join(tmpdir, f"{template_name}.pdf")
            try:
                subprocess.run(
                    [
                        "pdftocairo",
                        "-png",
                        "-singlefile",       # 단일 파일
                        "-cropbox",          # print area(cropbox) 기준으로 자름
                        "-scale-to-x", "2048",  # 가로 2048px 에 맞춰 스케일
                        "-f", "1", "-l", "1", # 1페이지만
                        pdf_path,
                        os.path.join(tmpdir, template_name)
                    ],
                    check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE
                )
            except subprocess.CalledProcessError as e:
                detail = e.stderr.decode(errors="ignore")
                raise HTTPException(500, detail=f"PDF→PNG 변환 실패: {detail}")
            
            # 4) 생성된 PNG 파일 찾기
            png_files = glob.glob(os.path.join(tmpdir, "*.png"))
            if not png_files:
                raise HTTPException(500, detail="변환된 PNG 파일을 찾을 수 없습니다.")
            png_path = png_files[0]

        # 5) 자동 크롭
        try:
            img = Image.open(png_path)
            bg = Image.new(img.mode, img.size, img.getpixel((0, 0)))
            diff = ImageChops.difference(img, bg)
            bbox = diff.getbbox()
            if bbox:
                img.crop(bbox).save(png_path)
        except Exception:
            # 크롭 실패해도 그냥 원본 내보내도록 무시
            print("자동 크롭 실패:", e)

        return png_path